from datetime import datetime, timedelta
from time import sleep
import base64
import os

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from bs4 import BeautifulSoup

from web_app.tools import fast_bitrix_client, send_bitrix_request, get_folder_id, get_user_folder_id


b = fast_bitrix_client()


def get_employee_id(users: str) -> list:
    """
    Приводит строку с id пользователей и id подразделений к единому списку, состоящему только из id сотрудников

    :param users: Строка из параметра запроса Б24 состоящая из {user_...} и|или {group_...}, которые разделены ', '
    """

    users_id_set = set()

    # Строка с сотрудниками и отделами в список
    users = users.split(', ')

    # Если в массиве найден id сотрудника
    for user_id in users:
        if 'user' in user_id:
            users_id_set.add(user_id[5:])

        # Если в массиве найден id отдела
        elif 'group' in user_id:
            department_users = b.get_all('user.get', {'filter': {'UF_DEPARTMENT': user_id[8:]}})
            for user in department_users:
                users_id_set.add(user['ID'])

    return list(users_id_set)


def change_sheet_style(sheet, add_filters=True, change_width=True, change_colors_and_fonts=True, change_fonts=False) -> None:
    """
    Добавляет фильтры в заголовки таблицы
    Изменяет ширину столбцов
    Перекрашивает ряды ячеек

    :param sheet: экземпляр класса Worksheet (выбранный в экземпляре класса Workbook) из openpyxl
    :param add_filters: добавить фильтры
    :param change_width: изменить ширину столбцов
    :param change_colors_and_fonts: изменить цвет ячеек и шрифт
    :param change_fonts: изменить шрифт
    """

    # Добавление фильтров в заголовки
    if add_filters:
        sheet.auto_filter.ref = sheet.dimensions

    # Изменение ширины
    cells_width = {
        0: 15,
        1: 15,
        2: 15,
        3: 10,
        4: 40,
        5: 7,
        6: 40,
        7: 10,
        8: 40,
        9: 10,
        10: 10
    }
    if change_width:
        for index, column_cells in enumerate(sheet.columns):
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = cells_width[index]
    else:
        for index, column_cells in enumerate(sheet.columns):
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = 50

    # Цвет ячеек
    if change_colors_and_fonts:
        for index, row_cells in enumerate(sheet.rows):
            if index == 0:
                cell_color = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                font_color = Font(color='FFFFFFFF', bold=True, name='Calibri', size=8)
                for cell in row_cells:
                    cell.fill = cell_color
                    cell.font = font_color

            elif index % 2 == 1:
                cell_color = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                font_color = Font(name='Calibri', size=8)
                for cell in row_cells:
                    cell.fill = cell_color
                    cell.font = font_color
            else:
                for cell in row_cells:
                    font_color = Font(name='Calibri', size=8)
                    cell.font = font_color

    # Автоперенос текста и выравнивание
    for index, column_rows in enumerate(sheet.rows):
        for cell in column_rows:
            cell.alignment = cell.alignment.copy(wrapText=True)
            if index == 0:
                cell.alignment = Alignment(horizontal='center')

    # Изменить шрифт
    if change_fonts:
        font_color = Font(name='Calibri', size=8)
        for row_cells in sheet.rows:
            for cell in row_cells:
                cell.font = font_color


def get_fio_from_user_info(user_info: dict) -> str:
    """
    Возвращает строку, состоящую из фамилии и имени пользователя Б24

    :param user_info: Словарь, полученный запросом с методом user.get
    """

    return f'{user_info["LAST_NAME"] if "LAST_NAME" in user_info else ""}'\
           f' {user_info["NAME"] if "NAME" in user_info else ""}'.strip()


def formate_iso_date(iso_date: str) -> str:
    """
    Преборазовывает дату iso формата в dd.mm.YYYY

    :param iso_date: Строка с датой в формате iso
    """

    if not iso_date:
        return ''

    try:
        return datetime.fromisoformat(iso_date).strftime('%d.%m.%Y')
    except ValueError:
        return ''


def get_company_and_title_from_activity(activity_info: dict) -> dict:
    """
    Получает ID компании из активности и название элемента CRM сущности.

    Если тип источника активности компания, тогда ID компании берется из "OWNER_ID"
    Если тип содержится в переменной "crm_types", тогда отправляется запрос crm.{тип активности}.get,
    из полученной информации берется значение COMPANY_ID
    В остальных случаях считается, что тип источника соответствует смарт-процессу и берется значение companyId из
    соответствуюшего элемента

    Отправляется запрос на получение информации о компании ID которой был найден и возвращается значение "TITLE"

    :param activity_info: словарь с информацией по активности, полученный запросом с методом "crm.activity.get|list"

    :return: Словарь с ключами 'Компания' со значением название компании, 'Название' с названием элемента,
            'ID компании', 'Ссылка на источник'
    """

    company_id = ''
    title = ''
    company_name = ''
    source_url = ''
    if activity_info['OWNER_TYPE_ID'] == '4':
        company_id = activity_info['OWNER_ID']
        source_url = f'https://avtograph.bitrix24.ru/crm/company/details/{activity_info["OWNER_ID"]}/'

    else:
        crm_types = {
            '1': 'lead',
            '2': 'deal',
            '3': 'contact',
            '7': 'quote',
            '5': 'invoice',
        }

        if activity_info['OWNER_TYPE_ID'] in crm_types:
            source_url = f'https://avtograph.bitrix24.ru/crm/{crm_types[activity_info["OWNER_TYPE_ID"]]}/details/{activity_info["OWNER_ID"]}/'
            crm_info = send_bitrix_request(f'crm.{crm_types[activity_info["OWNER_TYPE_ID"]]}.get', {
                'ID': activity_info['OWNER_ID']
            })

            if activity_info["OWNER_TYPE_ID"] == '3':
                title = get_fio_from_user_info(crm_info)
            else:
                title = crm_info['TITLE']

            if 'COMPANY_ID' in crm_info:
                company_id = crm_info['COMPANY_ID']

        else:
            smart_process_element = send_bitrix_request('crm.item.get', {
                'entityTypeId': activity_info['OWNER_TYPE_ID'],
                'id': activity_info['OWNER_ID']
            })

            if smart_process_element:
                smart_process_element = smart_process_element['item']
                title = smart_process_element['title']
                source_url = f'https://avtograph.bitrix24.ru/crm/type/{activity_info["OWNER_TYPE_ID"]}/details/{activity_info["OWNER_ID"]}/'
                if 'companyId' in smart_process_element:
                    company_id = smart_process_element['companyId']

    if company_id:
        company_info = send_bitrix_request('crm.company.get', {
            'ID': company_id
        })
        if company_info:
            company_name = company_info['TITLE']
            if activity_info['OWNER_TYPE_ID'] == '4':
                title = company_name

            return {
                'Компания': company_name,
                'Название': title,
                'ID компании': company_info['ID'],
                'Ссылка на источник': source_url
            }

    return {
        'Компания': company_name,
        'Название': title,
        'Ссылка на источник': source_url
    }


def create_activities_report(request: dict) -> None:
    """
    Запускается из БП "Отчет по активностям" в Новостях
    Создает отчет по активностям с заголовками из переменной "report_data"
    Отчет загружается на диск в Б24 в папку "Отчеты_по_активностям"
    Пользователь, запустивший БП, получает ссылку на файл с отчетом сообщением в уведомлениях

    :param request: словарь с параметрами запроса из Б24 с ключами:

                    date_start: Дата начала фильтра активностей
                    date_end: Дата конца фильтра активностей (может быть пустой строкой, тогда выбирается текущая дата)
                    who_starts: Пользователь, запустивший БП
                    users: Пользователи и отделы по которым нужно сформировать отчет
    """

    start_log_time = datetime.now()
    ddmmyyyy_pattern = '%d.%m.%Y'
    date_start = datetime.strptime(request['date_start'], ddmmyyyy_pattern)
    date_end = datetime.strptime(request['date_end'], ddmmyyyy_pattern) if request['date_end'] else datetime.now()
    activity_types = send_bitrix_request('crm.enum.activitytype')

    source_types = {
        '1': 'Лид',
        '2': 'Сделка',
        '3': 'Контакт',
        '4': 'Компания',
        '7': 'Предложение',
        '5': 'Счет (старый)',
        '8': 'Реквизиты',
        '12': 'Обзвон',
        '31': 'Счет',
    }

    # Дополнение словаря с типами crm смарт-процессами
    crm_types = send_bitrix_request('crm.type.list')['types']
    for crm_type in crm_types:
        source_types[crm_type['entityTypeId']] = crm_type['title']

    report_data = [
        [
            'Кто создал', 'Ответственный', 'Компания', 'Исполнить', 'Относится к', 'Тип активности', 'Тема',
            'Статус', 'Результат', 'Создано', 'П.ред.',
        ]
    ]

    activities = []
    users_id = get_employee_id(request['users']) if request['users'] else []
    while date_start.strftime(ddmmyyyy_pattern) != (date_end + timedelta(days=1)).strftime(ddmmyyyy_pattern):
        temp_activities = b.get_all('crm.activity.list', {
            'filter': {
                '>=CREATED': date_start.strftime(ddmmyyyy_pattern),
                '<CREATED': (date_start + timedelta(days=1)).strftime(ddmmyyyy_pattern),
                'RESPONSIBLE_ID': users_id,
            }
        })
        activities += temp_activities
        print(date_start)
        date_start = date_start + timedelta(days=1)
        sleep(1)

    users_info = b.get_all('user.get')

    for index, activity in enumerate(activities, 1):
        print(index, '|', len(activities))
        author_info = list(filter(lambda x: x['ID'] == activity['AUTHOR_ID'], users_info))[0]
        author_name = get_fio_from_user_info(author_info)

        responsible_info = list(filter(lambda x: x['ID'] == activity['RESPONSIBLE_ID'], users_info))[0]
        responsible_name = get_fio_from_user_info(responsible_info)

        if activity['PROVIDER_ID'] == 'CRM_TASKS_TASK':
            activity_type = 'Задача'
        else:
            activity_type = list(filter(lambda x: str(x['ID']) == activity['TYPE_ID'], activity_types))[0]['NAME']
            if activity_type == 'Пользовательское действие':
                activity_type = 'Дело'

        activity_source = source_types.get(activity['OWNER_TYPE_ID'], '')
        if not activity_source:
            activity_source = list(filter(lambda x:  str(x['entityTypeId']) == activity['OWNER_TYPE_ID'], crm_types))
            if activity_source:
                activity_source = activity_source[0]['title']
            else:
                activity_source = ''

        activity_result = ''
        if activity['TYPE_ID'] in ['6', '2']:
            note_information = send_bitrix_request('crm.timeline.note.get', {
                'ownerTypeId': activity['OWNER_TYPE_ID'],
                'ownerId': activity['OWNER_ID'],
                'itemType': '2' if activity['TYPE_ID'] in ['6', '2'] else '1',
                'itemId': activity['ID']
            })

            if note_information:
                activity_result = note_information['text']
        else:
            activity_result = activity['DESCRIPTION'] if 'DESCRIPTION' in activity else ''

        company_and_title = get_company_and_title_from_activity(activity)
        if company_and_title['Компания']:
            company_cell = f'{company_and_title["ID компании"]} {company_and_title["Компания"]}'
        else:
            company_cell = ''

        source_cell = f'{company_and_title["Ссылка на источник"]} {activity_source}: {company_and_title["Название"]}'

        end_time = formate_iso_date(activity['LAST_UPDATED']) if activity['COMPLETED'] == 'Y' else ''

        report_data.append(
            [
                author_name,                                                            # Кто создал
                responsible_name,                                                       # Ответственный
                company_cell,                                                           # Компания
                formate_iso_date(activity['DEADLINE']),                                 # Исполнить
                source_cell,                                                            # Относится к
                activity_type,                                                          # Тип активности
                activity['SUBJECT'],                                                    # Тема
                'Выполнено' if activity['COMPLETED'] == 'Y' else 'Не выполнено',        # Статус
                BeautifulSoup(activity_result, "lxml").text,                            # Результат
                formate_iso_date(activity['CREATED']),                                  # Создано
                end_time,                                                               # П.ред.
            ]
        )
        sleep(1)

    # Формирование файла
    report_name = f'Отчет_по_активностям_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'
    workbook = openpyxl.Workbook()

    # Первый лист "Отчет"
    sheet = workbook.active
    sheet.title = 'Отчет'
    for row in report_data:
        sheet.append(row)

    # Создание гиперссылок
    for row in range(2, len(report_data) + 1):
        company_cell = sheet['C' + str(row)]
        source_cell = sheet['E' + str(row)]
        if company_cell.value:
            company_cell_value = company_cell.value.split()
            company_cell.hyperlink = f"https://avtograph.bitrix24.ru/crm/company/details/{company_cell_value[0]}/"
            company_cell.value = ' '.join(company_cell_value[1:])
        source_cell_value = source_cell.value.split()
        if 'https' in source_cell_value[0]:
            source_cell.hyperlink = source_cell_value[0]
            source_cell.value = ' '.join(source_cell_value[1:])
    change_sheet_style(sheet)

    # Второй лист "Сводные"
    sheet = workbook.create_sheet('Сводные')
    sheet.append(['Тип активности', 'Количество', 'Выполнено', 'Не выполнено'])
    for activity_type in sorted(set(map(lambda x: x[5], report_data[1:]))):
        filtered_activities = list(filter(lambda x: x[5] == activity_type, report_data[1:]))
        sheet.append(
            [
                activity_type,
                len(filtered_activities),
                len(list(filter(lambda x: x[8] == 'Выполнено', filtered_activities))),
                len(list(filter(lambda x: x[8] != 'Выполнено', filtered_activities)))
            ]
        )
    change_sheet_style(sheet)

    # Третий лист "Инфо"
    sheet = workbook.create_sheet('Инфо')
    end_log_time = datetime.now()
    log_datetime_pattern = '%d.%m.%Y %H:%M:%S'
    info_data = [
        [
            'Отчет по активностям',
            f'с {request["date_start"]} по {date_end.strftime(ddmmyyyy_pattern)}',
        ],
        [
            f'Начало формирования: {start_log_time.strftime(log_datetime_pattern)}',
        ],
        [
            f'Конец формирования: {end_log_time.strftime(log_datetime_pattern)}',
        ],
        [
            f'Затраченное время: {(end_log_time - start_log_time)}',
        ],
        [''],
        [
            'Параметры БП:',
        ],
        [
            f'Дата начала (date_start): {request["date_start"]}',
        ],
        [
            f'Дата конца (date_end): {request["date_end"]}',
        ],
        [
            f'Пользователи (users): {request["users"]}',
        ],
        [
            f'Запущен (who_starts): {request["who_starts"]}',
        ],
    ]

    for info in info_data:
        sheet.append(info)
    change_sheet_style(sheet, add_filters=False, change_colors_and_fonts=False, change_width=False, change_fonts=True)

    workbook.save(report_name)

    # Загрузка отчета в Битрикс

    bitrix_folder_id = get_user_folder_id(request['who_starts'][5:])
    with open(report_name, 'rb') as file:
        report_file = file.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]
    upload_report = b.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })

    send_bitrix_request('im.notify.system.add', {
        'USER_ID': request['who_starts'][5:],
        'MESSAGE': f'Отчет по активностям сформирован. {upload_report["DETAIL_URL"]}'})
    os.remove(report_name)
